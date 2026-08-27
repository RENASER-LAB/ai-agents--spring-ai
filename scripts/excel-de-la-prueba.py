#!/usr/bin/env python3
"""
Vuelca a un Excel las notas de la prueba del puesto de una vacante.

POR QUE EXISTE
--------------
La decision de a quien se pasa a la siguiente etapa se toma mirando notas, y mirarlas de una
en una en el panel no deja compararlas. Esto saca la misma informacion que hay ahi, en dos
hojas que se pueden ordenar, filtrar y mandar a quien decide.

NO ESCRIBE NADA. Solo lee y guarda un .xlsx.

LAS DOS HOJAS
-------------
«Resumen»  una fila por candidato: como contactarlo —correo y telefono, los dos sacados del
           curriculum, no de la cuenta—, su nota sobre 100, cuantos criterios tiene puntuados
           y en que estado esta. La nota se calcula aqui a partir de la rubrica, igual que en
           `calificar-pruebas.py` y por el mismo motivo que se explica alli.

«Detalle»  una fila por criterio y candidato, con el puntaje, el maximo, quien lo puso
           (AGENTE o PERSONA) y la explicacion con la que lo justifico. Es el respaldo de la
           decision: sin la explicacion, la nota es un numero que nadie puede discutir.

«Respuestas»  una fila por pregunta y candidato, con lo que escribio. Sale tambien la que
           dejo en blanco, con la respuesta vacia: que alguien no contestara la cuarta es
           justo lo que hay que poder ver. Va en formato largo y no en veinte columnas
           porque una respuesta puede tener 20.000 caracteres.

UNA NOTA VACIA NO ES UN CERO
----------------------------
Sale vacia cuando la rubrica no esta entera —porque el agente no pudo puntuar algo, o porque
ese criterio es de metodo persona— y la columna «criterios» dice cuantos faltan. Un cero ahi
seria un juicio que nadie ha hecho.

USO
---
    python scripts/excel-de-la-prueba.py --vacante 14
    python scripts/excel-de-la-prueba.py --vacante 14 --salida /ruta/notas.xlsx
"""
import argparse
import json
import sys
import urllib.error
import urllib.request

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BACKEND = "https://18-204-177-210.nip.io"

# Los estados en que una prueba ya se entrego. Antes de eso no hay nada que volcar.
ENTREGADA = "PRUEBA_CALIFICANDO"
CALIFICADA = "PRUEBA_POR_CONFIRMAR"


def llamar(metodo, ruta, token=None, cuerpo=None):
    cuerpo_json = json.dumps(cuerpo).encode() if cuerpo is not None else None
    peticion = urllib.request.Request(BACKEND + ruta, data=cuerpo_json, method=metodo)
    if cuerpo_json is not None:
        peticion.add_header("Content-Type", "application/json")
    if token:
        peticion.add_header("Authorization", "Bearer " + token)
    try:
        with urllib.request.urlopen(peticion, timeout=120) as respuesta:
            texto = respuesta.read().decode()
            return respuesta.status, (json.loads(texto) if texto.strip() else None)
    except urllib.error.HTTPError as error:
        return error.code, error.read().decode()[:300]
    except urllib.error.URLError as error:
        return 0, str(error.reason)


def direccion(fila):
    """
    La direccion a la que se le puede escribir.

    NO es `correo`: esa es la de la cuenta, y en esta convocatoria se la invento el cargador
    («@cv-convocatoria.local», un dominio que no existe). La de verdad la saco el agente del
    propio curriculum y viaja dentro de `datos`.
    """
    return ((fila.get("datos") or {}).get("email") or "").strip()


def telefono(fila):
    """
    El telefono que el agente saco del curriculum, tal cual venia escrito.

    Se deja como texto y no se normaliza: llega en media docena de formatos —con prefijo,
    con guiones, con espacios— y unificarlos a ciegas es la clase de arreglo que convierte
    un numero raro pero correcto en uno que no existe. Quien vaya a llamar lo lee bien igual.
    """
    return ((fila.get("datos") or {}).get("telefono") or "").strip()


def nota_sobre_cien(notas):
    """La nota, o None si la rubrica no esta entera. Ver el modulo de calificar."""
    puestas = [n for n in notas if n.get("puntaje") is not None]
    maximo = sum(n["puntosMaximos"] for n in notas if n.get("puntosMaximos"))
    if not notas or len(puestas) != len(notas) or not maximo:
        return None
    return round(sum(n["puntaje"] for n in puestas) * 100.0 / maximo, 2)


CABECERA = PatternFill("solid", fgColor="4338CA")
BLANCA = Font(color="FFFFFF", bold=True)


def encabezar(hoja, titulos, anchos):
    hoja.append(titulos)
    for columna, ancho in enumerate(anchos, 1):
        hoja.column_dimensions[get_column_letter(columna)].width = ancho
    for celda in hoja[1]:
        celda.fill = CABECERA
        celda.font = BLANCA
        celda.alignment = Alignment(vertical="center")
    hoja.freeze_panes = "A2"


def main():
    opciones = argparse.ArgumentParser(description=__doc__,
                                       formatter_class=argparse.RawDescriptionHelpFormatter)
    opciones.add_argument("--vacante", type=int, required=True)
    opciones.add_argument("--usuario", default="andy-dev")
    opciones.add_argument("--salida", default=None)
    args = opciones.parse_args()

    codigo, sesion = llamar("POST", "/api/v1/panel/auth/dev-login", None,
                            {"usuarioRenaserOsId": args.usuario})
    if codigo != 200:
        sys.exit(f"No se pudo entrar como «{args.usuario}»: {sesion}")
    token = sesion["token"]

    codigo, ranking = llamar("GET", f"/api/v1/panel/vacantes/{args.vacante}/ranking", token)
    if codigo != 200:
        sys.exit(f"No se pudo leer la vacante {args.vacante}: {ranking}")

    filas = [f for f in ranking["filas"] if f["estado"] in (ENTREGADA, CALIFICADA)]
    if not filas:
        sys.exit("Nadie ha entregado todavia: no hay nada que volcar.")

    libro = Workbook()
    resumen = libro.active
    resumen.title = "Resumen"
    encabezar(resumen,
              ["#", "Candidato", "Correo", "Telefono", "Nota /100", "Criterios", "Estado",
               "Postulacion"],
              [5, 34, 38, 15, 11, 11, 24, 12])

    respuestas_hoja = libro.create_sheet("Respuestas")
    encabezar(respuestas_hoja,
              ["Candidato", "#", "Codigo", "Pregunta", "Respuesta"],
              [34, 5, 10, 60, 120])

    detalle = libro.create_sheet("Detalle")
    encabezar(detalle,
              ["Candidato", "Criterio", "Puntaje", "Maximo", "Puso", "Explicacion"],
              [34, 42, 9, 9, 10, 110])

    cuantas_con_nota, sin_responder, faltan_respuestas = 0, 0, []
    for numero, fila in enumerate(filas, 1):
        codigo, notas = llamar(
            "GET", f"/api/v1/panel/postulaciones/{fila['postulacionId']}/prueba/notas", token)
        notas = notas if codigo == 200 and isinstance(notas, list) else []
        puestas = [n for n in notas if n.get("puntaje") is not None]
        nota = nota_sobre_cien(notas)
        if nota is not None:
            cuantas_con_nota += 1

        resumen.append([
            numero,
            fila.get("candidato") or "",
            direccion(fila),
            telefono(fila),
            nota,                                     # vacia si la rubrica no esta entera
            f"{len(puestas)}/{len(notas)}" if notas else "sin prueba",
            fila.get("estadoNombre") or fila.get("estado") or "",
            fila["postulacionId"],
        ])

        codigo_r, suyas = llamar(
            "GET", f"/api/v1/panel/postulaciones/{fila['postulacionId']}/prueba/respuestas", token)
        if codigo_r == 200 and isinstance(suyas, list):
            sin_responder += sum(1 for r in suyas if not (r.get("respuesta") or "").strip())
            # Se numera por POSICION en la lista, no con `orden`: ese campo es el del
            # catalogo de preguntas y no el de esta plantilla —ADMIN_Q01 llega con un 7—.
            # La lista si viene ordenada como la vio el candidato, que es lo que vale.
            for posicion, r in enumerate(suyas, 1):
                respuestas_hoja.append([
                    fila.get("candidato") or "",
                    posicion,
                    r.get("codigo") or "",
                    (r.get("enunciado") or "").strip(),
                    (r.get("respuesta") or "").strip(),
                ])
        elif codigo_r == 404:
            faltan_respuestas.append(fila.get("candidato") or "")
        else:
            faltan_respuestas.append(f"{fila.get('candidato')} (HTTP {codigo_r})")

        for n in notas:
            detalle.append([
                fila.get("candidato") or "",
                n.get("nombre") or "",
                n.get("puntaje"),
                n.get("puntosMaximos"),
                n.get("origen") or "",
                (n.get("explicacion") or "").strip(),
            ])

    for celda in detalle["F"][1:]:
        celda.alignment = Alignment(wrap_text=True, vertical="top")
    for celda in respuestas_hoja["D"][1:]:
        celda.alignment = Alignment(wrap_text=True, vertical="top")
    for celda in respuestas_hoja["E"][1:]:
        celda.alignment = Alignment(vertical="top")

    salida = args.salida or f"notas-prueba-vacante-{args.vacante}.xlsx"
    libro.save(salida)
    print(f"{len(filas)} candidatos · {cuantas_con_nota} con la rubrica entera")
    print(f"{respuestas_hoja.max_row - 1} respuestas · {sin_responder} preguntas en blanco")
    if faltan_respuestas:
        # Se dice, no se calla: una hoja a la que le faltan candidatos sin avisar se lee
        # como si esos no hubieran contestado nada.
        print(f"SIN respuestas ({len(faltan_respuestas)}): " + ", ".join(faltan_respuestas))
    print(f"Escrito: {salida}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(130)
