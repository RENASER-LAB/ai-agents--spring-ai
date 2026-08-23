# Carga los 3 archivos Excel del banco v3 en la base local como bancos nuevos publicados,
# y luego los compara texto a texto contra el banco original de las migraciones.
# Es el ensayo del flujo que formalizará el importador del panel: Excel -> base -> portal.
import re
import subprocess
import sys
from openpyxl import load_workbook

BASE = '/home/n4nd0/Documentos/RENASER-RECLUTAMIENTO/docs/insumos/'
ROLES = [
    ('banco-v3-directivo.xlsx', 'DIRECCION', 'Banco desde Excel · Directivo — prueba local',
     'Banco RENASER v3 · Directivo'),
    ('banco-v3-coordinacion-y-supervision.xlsx', 'SUPERVISION',
     'Banco desde Excel · Coordinación y Supervisión — prueba local',
     'Banco RENASER v3 · Coordinación y Supervisión'),
    ('banco-v3-ejecutivo-y-operativo.xlsx', 'EJECUCION',
     'Banco desde Excel · Ejecutivo y Operativo — prueba local',
     'Banco RENASER v3 · Ejecutivo y Operativo'),
]
LETRAS = 'abcdefghijklmnopqrstuvwxyz'


def psql(consulta, entrada=None):
    orden = ['docker', 'exec', '-i', 'renaser-postgres', 'psql', '-U', 'postgres',
             '-d', 'renaser_db', '-v', 'ON_ERROR_STOP=1', '-At', '-F', '\x01', '-R', '\x02']
    if entrada is None:
        orden += ['-c', consulta]
    r = subprocess.run(orden, input=entrada, capture_output=True, text=True)
    if r.returncode != 0:
        print(r.stderr, file=sys.stderr)
        sys.exit(1)
    return [[c.strip('\n') for c in f.split('\x01')]
            for f in r.stdout.split('\x02') if f.strip()]


def lit(x):
    if x is None or x == '':
        return 'NULL'
    return "'" + str(x).replace("'", "''") + "'"


def norm(s):
    return re.sub(r'\s+', ' ', (s or '')).strip()


def filas(ws):
    for fila in ws.iter_rows(min_row=5, values_only=True):
        if fila[0] is None:
            break
        yield fila


def cargar(archivo, nivel, etiqueta):
    wb = load_workbook(BASE + archivo, read_only=True)
    sql = ['BEGIN;']
    sql.append(f"""
INSERT INTO version_banco (organizacion_id, tipo_banco, nivel_puesto_codigo, etiqueta, estado, publicada_en)
SELECT id, 'NIVEL', {lit(nivel)}, {lit(etiqueta)}, 'PUBLICADA', now()
  FROM organizacion WHERE codigo = 'RENASER';""")

    for orden, f in enumerate(filas(wb['Preguntas']), start=1):
        codigo, tipo, enunciado, situacion, peso, elim, _mide, n_campos, nota = f[:9]
        peso = int(peso or 0)
        sql.append(f"""
INSERT INTO pregunta (version_banco_id, codigo, tipo, enunciado, situacion, logica_interna,
                      es_puntuable, orden, peso, es_clave, es_eliminatorio, casos_pedidos)
SELECT vb.id, {lit(codigo)}, {lit(tipo)}, {lit(enunciado)}, {lit(situacion)}, {lit(nota)},
       {'true' if peso > 0 else 'false'}, {orden}, {peso}, false,
       {'true' if elim == 'sí' else 'false'}, {int(n_campos) if n_campos else 'NULL'}
  FROM version_banco vb WHERE vb.etiqueta = {lit(etiqueta)};""")

    contador = {}
    for f in filas(wb['Opciones']):
        codigo, texto, puntaje, valor, trampa, pos = f[:6]
        letra = LETRAS[contador.setdefault(codigo, 0)]
        contador[codigo] += 1
        sql.append(f"""
INSERT INTO opcion (pregunta_id, letra, texto, puntaje, valor, es_distractor, orden_correcto)
SELECT p.id, {lit(letra)}, {lit(texto)}, {int(puntaje) if puntaje else 'NULL'},
       {int(valor) if valor not in (None, '') else 'NULL'},
       {'true' if trampa == 'sí' else 'false'}, {int(pos) if pos else 'NULL'}
  FROM pregunta p JOIN version_banco vb ON vb.id = p.version_banco_id
 WHERE vb.etiqueta = {lit(etiqueta)} AND p.codigo = {lit(codigo)};""")

    contador = {}
    for f in filas(wb['Campos de caso (CD)']):
        codigo, campo, validacion = f[:3]
        orden = contador.setdefault(codigo, 0) + 1
        contador[codigo] = orden
        sql.append(f"""
INSERT INTO campo_caso (pregunta_id, orden, etiqueta, validacion)
SELECT p.id, {orden}, {lit(campo)}, {lit(validacion)}
  FROM pregunta p JOIN version_banco vb ON vb.id = p.version_banco_id
 WHERE vb.etiqueta = {lit(etiqueta)} AND p.codigo = {lit(codigo)};""")

    contador = {}
    for f in filas(wb['Rangos (V)']):
        codigo, condicion, puntos, bandera = f[:4]
        orden = contador.setdefault(codigo, 0) + 1
        contador[codigo] = orden
        sql.append(f"""
INSERT INTO rango_pregunta (pregunta_id, orden, condicion, puntaje, genera_bandera)
SELECT p.id, {orden}, {lit(condicion)}, {int(puntos)}, {'true' if bandera == 'sí' else 'false'}
  FROM pregunta p JOIN version_banco vb ON vb.id = p.version_banco_id
 WHERE vb.etiqueta = {lit(etiqueta)} AND p.codigo = {lit(codigo)};""")

    for f in filas(wb['Pares']):
        a, b, pen, sep, cond = f[:5]
        sql.append(f"""
INSERT INTO par_consistencia (version_banco_id, pregunta_a_id, pregunta_b_id,
                              penalizacion_porcentaje, separacion_minima_items, condicion)
SELECT vb.id, pa.id, pb.id, {pen}, {sep}, {lit(cond)}
  FROM version_banco vb
  JOIN pregunta pa ON pa.version_banco_id = vb.id AND pa.codigo = {lit(a)}
  JOIN pregunta pb ON pb.version_banco_id = vb.id AND pb.codigo = {lit(b)}
 WHERE vb.etiqueta = {lit(etiqueta)};""")

    sql.append('COMMIT;')
    psql(None, entrada='\n'.join(sql))


def traer(etiqueta):
    vb = f"(select id from version_banco where etiqueta = {lit(etiqueta)})"
    datos = {}
    datos['preguntas'] = {
        f[0]: tuple(norm(x) for x in f[1:])
        for f in psql(f"""select codigo, tipo, enunciado, coalesce(situacion,''),
                          peso::int::text, es_eliminatorio::text, es_puntuable::text,
                          coalesce(casos_pedidos::text,''), coalesce(logica_interna,'')
                          from pregunta where version_banco_id = {vb}""")}
    datos['opciones'] = {}
    for f in psql(f"""select p.codigo, o.texto, coalesce(o.puntaje::text,''),
                      coalesce(o.valor::text,''), o.es_distractor::text,
                      coalesce(o.orden_correcto::text,'')
                      from opcion o join pregunta p on p.id = o.pregunta_id
                      where p.version_banco_id = {vb} order by p.orden, o.id"""):
        datos['opciones'].setdefault(f[0], []).append(tuple(norm(x) for x in f[1:]))
    datos['campos'] = {
        (f[0], f[1]): (norm(f[2]), norm(f[3]))
        for f in psql(f"""select p.codigo, cc.orden::text, cc.etiqueta, coalesce(cc.validacion,'')
                          from campo_caso cc join pregunta p on p.id = cc.pregunta_id
                          where p.version_banco_id = {vb}""")}
    datos['rangos'] = {
        (f[0], f[1]): (norm(f[2]), f[3], f[4])
        for f in psql(f"""select p.codigo, r.orden::text, r.condicion, r.puntaje::text,
                          r.genera_bandera::text
                          from rango_pregunta r join pregunta p on p.id = r.pregunta_id
                          where p.version_banco_id = {vb}""")}
    datos['pares'] = {
        (f[0], f[1]): (f[2], f[3], norm(f[4]))
        for f in psql(f"""select pa.codigo, pb.codigo, pc.penalizacion_porcentaje::text,
                          pc.separacion_minima_items::text, coalesce(pc.condicion,'')
                          from par_consistencia pc
                          join pregunta pa on pa.id = pc.pregunta_a_id
                          join pregunta pb on pb.id = pc.pregunta_b_id
                          where pc.version_banco_id = {vb}""")}
    return datos


def verificar(etiqueta_excel, etiqueta_original):
    excel, original = traer(etiqueta_excel), traer(etiqueta_original)
    fallos = []

    for codigo, fila in original['preguntas'].items():
        suya = excel['preguntas'].get(codigo)
        if suya is None:
            fallos.append(f'pregunta {codigo}: falta en el Excel')
        elif suya != fila:
            for i, (a, b) in enumerate(zip(fila, suya)):
                if a != b:
                    fallos.append(f'pregunta {codigo} campo {i}: «{a[:60]}» vs «{b[:60]}»')

    for codigo, lista in original['opciones'].items():
        suya = excel['opciones'].get(codigo, [])
        if len(lista) != len(suya):
            fallos.append(f'opciones de {codigo}: {len(lista)} vs {len(suya)}')
            continue
        for i, (a, b) in enumerate(zip(lista, suya)):
            if a != b:
                fallos.append(f'opción {codigo}#{i + 1}: {a} vs {b}')

    for clave, valor in original['campos'].items():
        if excel['campos'].get(clave) != valor:
            fallos.append(f'campo {clave}: «{valor[0][:60]}» vs '
                          f'«{(excel["campos"].get(clave) or ("(falta)",))[0][:60]}»')

    # Los rangos del Excel incluyen las copias de las 4 preguntas que comparten tabla
    # (rangos_de_pregunta_codigo); en el original esas no tienen filas propias.
    for clave, valor in original['rangos'].items():
        if excel['rangos'].get(clave) != valor:
            fallos.append(f'rango {clave}: {valor} vs {excel["rangos"].get(clave)}')
    extra = set(excel['rangos']) - set(original['rangos'])
    compartidas = {'C36', 'O32', 'O48', 'C54'}
    raras = {c for c, _ in extra} - compartidas
    if raras:
        fallos.append(f'rangos inesperados en el Excel: {sorted(raras)}')

    if original['pares'] != excel['pares']:
        fallos.append(f'pares: {original["pares"]} vs {excel["pares"]}')

    return fallos


if __name__ == '__main__':
    ya = psql("select etiqueta from version_banco where etiqueta like 'Banco desde Excel%'")
    if ya:
        usadas = psql("""select count(*) from evaluacion e join version_banco vb
                         on vb.id = e.version_banco_nivel_id
                         where vb.etiqueta like 'Banco desde Excel%'""")[0][0]
        if usadas != '0':
            sys.exit('Hay evaluaciones colgando de los bancos de prueba: bórralas primero.')
        psql(None, entrada="""BEGIN;
DELETE FROM par_consistencia WHERE version_banco_id IN
  (SELECT id FROM version_banco WHERE etiqueta LIKE 'Banco desde Excel%');
DELETE FROM rango_pregunta WHERE pregunta_id IN (SELECT p.id FROM pregunta p
  JOIN version_banco vb ON vb.id = p.version_banco_id WHERE vb.etiqueta LIKE 'Banco desde Excel%');
DELETE FROM campo_caso WHERE pregunta_id IN (SELECT p.id FROM pregunta p
  JOIN version_banco vb ON vb.id = p.version_banco_id WHERE vb.etiqueta LIKE 'Banco desde Excel%');
DELETE FROM opcion WHERE pregunta_id IN (SELECT p.id FROM pregunta p
  JOIN version_banco vb ON vb.id = p.version_banco_id WHERE vb.etiqueta LIKE 'Banco desde Excel%');
DELETE FROM pregunta WHERE version_banco_id IN
  (SELECT id FROM version_banco WHERE etiqueta LIKE 'Banco desde Excel%');
DELETE FROM version_banco WHERE etiqueta LIKE 'Banco desde Excel%';
COMMIT;""")
        print('Bancos de prueba anteriores retirados.')

    todo_bien = True
    for archivo, nivel, etiqueta, original in ROLES:
        cargar(archivo, nivel, etiqueta)
        fallos = verificar(etiqueta, original)
        n = psql(f"""select count(*) from pregunta p join version_banco vb
                     on vb.id = p.version_banco_id where vb.etiqueta = {lit(etiqueta)}""")[0][0]
        if fallos:
            todo_bien = False
            print(f'✗ {nivel}: {n} preguntas cargadas, {len(fallos)} diferencias:')
            for f in fallos[:15]:
                print('   ', f)
        else:
            print(f'✓ {nivel}: {n} preguntas cargadas desde {archivo}, idénticas al original.')
    sys.exit(0 if todo_bien else 1)
