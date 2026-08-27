#!/usr/bin/env python3
"""Revisa que los tres Excel de CAZATALENTOS esten completos para importarse.

NO ESCRIBE NADA. Solo lee y enseña lo que falta.

QUE COMPRUEBA
-------------
Por cada pregunta que puntua (peso 1 o 2): que tenga enunciado, pilar, C3 esperado,
C4 esperado y senal de 0, y que el peso y la marca de eliminatoria sean valores validos.
Las de cierre (Z01..Z03, peso 0) no llevan C3 ni C4 y por eso se cuentan aparte.

Ademas cruza los pilares que usan las preguntas contra los que declara la hoja «Calculo»:
un pilar con preguntas y sin peso no se puede calificar, y un peso sin preguntas es basura.

    python3 scripts/revisa-excel-cazatalentos.py
"""

import glob

import openpyxl

PESOS_VALIDOS = (0, 1, 2, '0', '1', '2')


def revisa(ruta):
    libro = openpyxl.load_workbook(ruta, data_only=True)
    hoja = libro['Prueba RENASER']
    filas = [f for f in hoja.iter_rows(min_row=5, values_only=True) if f[0]]

    puntuables = [f for f in filas if f[6] not in (0, '0')]
    cierre = [f for f in filas if f[6] in (0, '0')]

    problemas = []
    for cod, pilar, enunciado, c3, c4, senal, peso, elim, _nota in filas:
        if not enunciado:
            problemas.append('%s sin enunciado' % cod)
        if not pilar:
            problemas.append('%s sin pilar' % cod)
        if peso not in PESOS_VALIDOS:
            problemas.append('%s con peso invalido: %r' % (cod, peso))
        if elim not in ('sí', 'no'):
            problemas.append('%s con eliminatoria invalida: %r' % (cod, elim))
        if peso in (0, '0'):
            continue                       # las de cierre no puntuan: no llevan C3 ni C4
        for nombre, valor in (('C3', c3), ('C4', c4), ('senal de 0', senal)):
            if not valor:
                problemas.append('%s sin %s' % (cod, nombre))

    cuenta = {}
    for fila in puntuables:
        cuenta[fila[1]] = cuenta.get(fila[1], 0) + 1

    pesos = {f[0]: f[1] for f in libro['Calculo' if 'Calculo' in libro.sheetnames else 'Cálculo']
             .iter_rows(min_row=5, max_row=11, values_only=True) if f[0]}

    sin_peso = [p for p in cuenta if p not in pesos]
    sin_preguntas = [p for p in pesos if p not in cuenta and 'Integridad' not in p]

    terminos = sum(1 for f in libro['Textura'].iter_rows(min_row=5, values_only=True)
                   if f[0] and f[1] and str(f[0]).startswith('F'))

    print('== %s' % ruta.split('/')[-1])
    print('   puntuables: %d   ·   cierre: %d   ·   de peso 2: %d'
          % (len(puntuables), len(cierre),
             sum(1 for f in puntuables if f[6] in (2, '2'))))
    print('   pilares: %s' % ' · '.join('%s=%d' % (k.split()[0], v)
                                        for k, v in sorted(cuenta.items())))
    if sin_peso:
        print('   [!] pilar con preguntas y sin peso: %s' % sin_peso)
    if sin_preguntas:
        print('   [!] peso declarado sin preguntas: %s' % sin_preguntas)
    print('   terminos de textura: %d' % terminos)
    print('   problemas: %s' % ('; '.join(problemas) if problemas else 'ninguno'))
    return problemas or sin_peso or sin_preguntas


def main():
    rutas = sorted(glob.glob('docs/insumos/CAZATALENTOS-???.xlsx'))
    if not rutas:
        print('No se encontro ningun CAZATALENTOS-*.xlsx en docs/insumos/')
        return 1
    hubo = False
    for ruta in rutas:
        hubo = revisa(ruta) or hubo
    print()
    print('LISTOS PARA IMPORTAR' if not hubo else 'HAY QUE ARREGLAR LO DE ARRIBA')
    return 1 if hubo else 0


if __name__ == '__main__':
    raise SystemExit(main())
