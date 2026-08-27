#!/usr/bin/env python3
"""Compara los bancos CAZATALENTOS de la base contra los Excel de la clienta, texto a texto.

Existe por la misma razon que comparar-banco-v3-con-base.py: las comprobaciones del
importador cuentan y esta lee. Con el v3 llegaron a produccion enunciados cortados sin que
ningun total dejara de cuadrar; la unica forma de ver ese fallo es comparar lo guardado
contra la fuente.

    python3 scripts/comparar-cazatalentos-con-base.py                 # las últimas versiones
    python3 scripts/comparar-cazatalentos-con-base.py --version 12    # una versión concreta

Que compara: por cada nivel toma la version_banco de metodo CRITERIOS mas reciente (o la
pedida con --version) y exige igualdad campo a campo contra la hoja «Prueba RENASER» del
Excel de su nivel: codigo, enunciado, C3 esperado, C4 esperado, señal de 0, peso,
eliminatoria, orden y pilar (via pregunta_dimension). Sin excepciones: estos Excel no
tienen los cortes de página del PDF, así que todo lo que difiera es un fallo.

No toca la base: solo lee, via psql dentro del contenedor de docker-compose.
"""

import argparse
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parent.parent
INSUMOS = RAIZ / "docs" / "insumos"

BANCOS = [
    ("CAZATALENTOS-DIR.xlsx", "DIRECCION"),
    ("CAZATALENTOS-SUP.xlsx", "SUPERVISION"),
    ("CAZATALENTOS-OPE.xlsx", "EJECUCION"),
]

PILARES = {
    "1": "PIL_INICIATIVA", "2": "PIL_RESOLUCION", "3": "PIL_EXCELENCIA",
    "4": "PIL_SERVICIO", "5": "PIL_RESPONSABILIDAD", "6": "PIL_DIRECCION",
    "7": "PIL_INTEGRIDAD",
}

SEPARADOR_CAMPO = "\x01"
SEPARADOR_FILA = "\x02"


def psql(consulta, db="renaser_db"):
    orden = ["docker", "exec", "-i", "renaser-postgres", "psql", "-U", "postgres",
             "-d", db, "-v", "ON_ERROR_STOP=1", "-At",
             "-F", SEPARADOR_CAMPO, "-R", SEPARADOR_FILA]
    r = subprocess.run(orden, input=consulta, capture_output=True, text=True)
    if r.returncode != 0:
        sys.exit(r.stderr)
    texto = r.stdout.strip(SEPARADOR_FILA + "\n")
    if not texto:
        return []
    return [fila.split(SEPARADOR_CAMPO) for fila in texto.split(SEPARADOR_FILA)]


def blancos(texto):
    """Solo se normalizan los espacios: el contenido se exige idéntico."""
    return " ".join((texto or "").split())


def del_excel(ruta):
    hoja = load_workbook(ruta, data_only=True)["Prueba RENASER"]
    filas = list(hoja.iter_rows(values_only=True))
    encabezado = next(i for i, f in enumerate(filas) if f[0] == "Código")
    preguntas = []
    for f in filas[encabezado + 2:]:
        if not f[0]:
            continue
        codigo, pilar, enunciado, c3, c4, senal, peso, elim = f[:8]
        preguntas.append({
            "codigo": codigo,
            "pilar": PILARES.get(str(pilar).split()[0]) if pilar and int(peso) > 0 else None,
            "enunciado": blancos(enunciado),
            "c3": blancos(c3), "c4": blancos(c4), "senal": blancos(senal),
            "peso": str(peso), "eliminatoria": elim == "sí",
        })
    return preguntas


def de_la_base(version_id, db):
    filas = psql(f"""
        SELECT p.codigo, p.enunciado, coalesce(p.c3_esperado, ''),
               coalesce(p.c4_esperado, ''), coalesce(p.senal_de_cero, ''),
               p.peso::text, p.es_eliminatorio::text, p.orden::text,
               coalesce(pd.dimension_codigo, '')
          FROM pregunta p
          LEFT JOIN pregunta_dimension pd ON pd.pregunta_id = p.id
         WHERE p.version_banco_id = {int(version_id)}
         ORDER BY p.orden;""", db)
    return [{
        "codigo": f[0], "enunciado": blancos(f[1]), "c3": blancos(f[2]),
        "c4": blancos(f[3]), "senal": blancos(f[4]), "peso": f[5],
        # ::text de un boolean da 'true'/'false' (el 't' pelado es solo del formato crudo de psql)
        "eliminatoria": f[6] in ("t", "true"), "orden": int(f[7]), "pilar": f[8] or None,
    } for f in filas]


def comparar(nivel, ruta_excel, version_id, db):
    esperadas = del_excel(ruta_excel)
    guardadas = de_la_base(version_id, db)
    fallos = []

    if [e["codigo"] for e in esperadas] != [g["codigo"] for g in guardadas]:
        fallos.append("el orden o la lista de códigos no coincide:\n"
                      f"    excel: {[e['codigo'] for e in esperadas]}\n"
                      f"    base:  {[g['codigo'] for g in guardadas]}")
        return fallos

    for e, g in zip(esperadas, guardadas):
        for campo, nombre in (("enunciado", "enunciado"), ("c3", "C3 esperado"),
                              ("c4", "C4 esperado"), ("senal", "señal de 0"),
                              ("peso", "peso"), ("eliminatoria", "eliminatoria"),
                              ("pilar", "pilar")):
            if e[campo] != g[campo]:
                fallos.append(f"{e['codigo']} · {nombre}:\n"
                              f"    excel: {e[campo]!r}\n    base:  {g[campo]!r}")
    return fallos


def main():
    opciones = argparse.ArgumentParser(description=__doc__)
    opciones.add_argument("--version", type=int,
                          help="comparar esta version_banco (si no, la última de cada nivel)")
    opciones.add_argument("--db", default="renaser_db")
    args = opciones.parse_args()

    excel_por_nivel = {nivel: archivo for archivo, nivel in BANCOS}

    if args.version:
        # El Excel correcto lo dice la propia version: compararla contra el de otro
        # nivel reportaria diferencias falsas en todo.
        filas = psql(f"""
            SELECT nivel_puesto_codigo FROM version_banco
             WHERE id = {args.version} AND metodo_calificacion = 'CRITERIOS';""", args.db)
        if not filas:
            sys.exit(f"La versión {args.version} no existe o no es un banco CRITERIOS")
        nivel = filas[0][0]
        pendientes = [(nivel, args.version)]
    else:
        pendientes = []
        for archivo, nivel in BANCOS:
            filas = psql(f"""
                SELECT id FROM version_banco
                 WHERE metodo_calificacion = 'CRITERIOS' AND nivel_puesto_codigo = '{nivel}'
                 ORDER BY id DESC LIMIT 1;""", args.db)
            if not filas:
                print(f"== {nivel}: no hay ningún banco CRITERIOS importado — se salta")
                continue
            pendientes.append((nivel, filas[0][0]))

    hubo = False
    for nivel, version_id in pendientes:
        fallos = comparar(nivel, INSUMOS / excel_por_nivel[nivel], version_id, args.db)
        print(f"== {nivel} · versión {version_id} · "
              + ("TODO CASA" if not fallos else f"{len(fallos)} DIFERENCIAS"))
        for f in fallos:
            print("  -", f)
        hubo = hubo or bool(fallos)

    return 1 if hubo else 0


if __name__ == "__main__":
    raise SystemExit(main())
