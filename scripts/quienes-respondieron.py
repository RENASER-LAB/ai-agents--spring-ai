#!/usr/bin/env python3
"""
Los que ya entregaron su evaluacion, con su nota y por que quedaron donde quedaron.

POR QUE ESTA SEPARADO DE lista-para-rrhh.py
-------------------------------------------
Aquella lista es para llamar por telefono a una tanda: sale todo el ranking, respondan o no.
Esta contesta otra pregunta —«de los que ya contestaron, a quien hago pasar»— y por eso trae
las cuatro notas del Perfil de Talento y no solo el total.

LO QUE ESTA HOJA NO DICE
------------------------
Quien aprueba. El sistema no aprueba ni suspende: ordena y agrupa, y quien decide es una
persona. La columna «grupo» es esa agrupacion, y los umbrales que la producen (80 y 65) son
un valor puesto por nosotros que Renaser todavia no ha confirmado.

Por eso la hoja trae «confianza» al lado de la nota. Una nota alta con confianza baja
significa que la IA vio poco material —un curriculum que no se pudo leer, respuestas muy
cortas— y ahi el numero vale menos que en el resto de la columna.

USO
---
    python scripts/quienes-respondieron.py
    python scripts/quienes-respondieron.py --salida "C:/ruta/archivo.xlsx"
"""
import argparse
import json
import sys
import urllib.error
import urllib.request

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BACKEND = "https://18-204-177-210.nip.io"
VACANTES = [(15, "Arquitecto"), (16, "Ingeniero Civil")]

COLOR = {
    "ALTA": "2F6E51",
    "POTENCIAL_CON_RIESGO": "9B6A22",
    "NO_PRIORIZADO": "70706B",
    "INCOMPATIBLE": "A43B36",
}
LEGIBLE = {
    "ALTA": "Alta",
    "POTENCIAL_CON_RIESGO": "Potencial con riesgo",
    "NO_PRIORIZADO": "No priorizado",
    "INCOMPATIBLE": "Incompatible",
}

COLUMNAS = [
    ("#", 5), ("Candidato", 32), ("Nota", 8), ("Grupo", 21),
    ("Adecuación", 12), ("Potencial", 11), ("Confianza", 11),
    ("Teléfono", 15), ("Correo", 32),
    ("Riesgos", 9), ("Alertas", 9),
    ("Qué vio la IA", 90), ("ID", 7),
]


def llamar(metodo, ruta, token=None, cuerpo=None):
    datos = json.dumps(cuerpo).encode() if cuerpo is not None else None
    peticion = urllib.request.Request(BACKEND + ruta, data=datos, method=metodo)
    if datos is not None:
        peticion.add_header("Content-Type", "application/json")
    if token:
        peticion.add_header("Authorization", "Bearer " + token)
    try:
        with urllib.request.urlopen(peticion, timeout=120) as respuesta:
            texto = respuesta.read().decode()
            return respuesta.status, (json.loads(texto) if texto.strip() else None)
    except urllib.error.HTTPError as error:
        return error.code, error.read().decode()[:300]


def hoja_de(libro, titulo, filas):
    h = libro.create_sheet(titulo)
    for indice, (nombre, ancho) in enumerate(COLUMNAS, 1):
        celda = h.cell(row=1, column=indice, value=nombre)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="0D0D0D")
        h.column_dimensions[get_column_letter(indice)].width = ancho
    h.freeze_panes = "C2"

    for numero, fila in enumerate(filas, 1):
        datos = fila.get("datos") or {}
        grupo = fila.get("grupoPrioridad")
        valores = [
            numero,
            fila.get("candidato") or datos.get("nombre") or "",
            fila.get("notaEtapa"),
            LEGIBLE.get(grupo, grupo or ""),
            fila.get("adecuacion"),
            fila.get("potencial"),
            fila.get("confianzaEvidencia"),
            str(datos.get("telefono") or ""),
            datos.get("email") or "",
            fila.get("riesgosCriticos"),
            fila.get("alertas"),
            (fila.get("resumen") or "").strip(),
            fila.get("postulacionId"),
        ]
        for indice, valor in enumerate(valores, 1):
            celda = h.cell(row=numero + 1, column=indice, value=valor)
            if indice == 4 and grupo in COLOR:
                celda.font = Font(color=COLOR[grupo], bold=True)
            if indice == 12:
                celda.alignment = Alignment(wrap_text=True, vertical="top")
        h.row_dimensions[numero + 1].height = 58

        # Una nota alta con poca evidencia detras no vale lo mismo. Se marca para que quien
        # lea la hoja no compare dos numeros que no son comparables.
        confianza = fila.get("confianzaEvidencia")
        if isinstance(confianza, (int, float)) and confianza < 60:
            h.cell(row=numero + 1, column=7).fill = PatternFill("solid", fgColor="FBE9E7")
    return h


def main():
    opciones = argparse.ArgumentParser(description=__doc__,
                                       formatter_class=argparse.RawDescriptionHelpFormatter)
    opciones.add_argument("--usuario", default="andy-dev")
    opciones.add_argument("--salida", default="respondieron-el-examen.xlsx")
    opciones.add_argument("--solo", default="",
                          help="ids de postulacion separados por coma: solo esos")
    args = opciones.parse_args()
    solo = {int(x) for x in args.solo.split(",") if x.strip()}
    if not solo:
        print("AVISO: sin --solo salen tambien los que nunca respondieron el examen.",
              file=sys.stderr)

    estado, sesion = llamar("POST", "/api/v1/panel/auth/dev-login",
                            cuerpo={"usuarioRenaserOsId": args.usuario})
    if estado != 200:
        print(f"No se pudo entrar al panel: {estado} {sesion}", file=sys.stderr)
        return 1
    token = sesion["token"]

    libro = Workbook()
    libro.remove(libro.active)
    total = 0

    for vacante_id, titulo in VACANTES:
        estado, ranking = llamar("GET", f"/api/v1/panel/vacantes/{vacante_id}/ranking", token)
        if estado != 200:
            print(f"Vacante {vacante_id}: {estado} {ranking}", file=sys.stderr)
            return 1
        filas = ranking["filas"] if isinstance(ranking, dict) and "filas" in ranking else ranking

        # Los que de verdad ENTREGARON el examen.
        #
        # No basta con mirar el ranking: ahi tambien salen los que la IA califico solo con el
        # curriculum, porque la criba los adelanto sin que nadie respondiera nada. Los dos
        # llegan a PERFIL_POR_CONFIRMAR y en la tabla se ven identicos, pero uno tiene detras
        # un examen de 50 preguntas y el otro no. Mezclarlos seria comparar dos cosas
        # distintas justo cuando hay que decidir a quien se llama.
        #
        # Quien lo sabe es la evaluacion, y el ranking no la expone, asi que los ids se
        # pasan por fuera con --solo. Si no se pasa ninguno, sale el ranking entero y se
        # avisa de que ahi dentro hay de los dos tipos.
        if solo:
            suyos = [f for f in filas if f.get("postulacionId") in solo]
        else:
            suyos = [f for f in filas
                     if f.get("estado") == "PERFIL_POR_CONFIRMAR"
                     and f.get("estadoCalificacion") == "TERMINADA"]
        hoja_de(libro, titulo, suyos)
        total += len(suyos)
        print(f"{titulo}: {len(suyos)} respondieron")

    libro.save(args.salida)
    print(f"\n{total} en total. Guardado en {args.salida}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
