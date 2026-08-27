#!/usr/bin/env python3
"""Ajusta el alcance de R14 y R15 en el banco SUP de CAZATALENTOS.

POR QUE EXISTE
--------------
El .md de la clienta (parte 4) pide que para SUP se diga «tu equipo» en vez de «tu area»
en R14 y R15. Pero ninguna de las dos preguntas contiene la palabra «area»: la instruccion
esta escrita contra un borrador anterior del cuestionario.

Se aplica su INTENCION declarada —«las mismas preguntas del nivel DIR, ajustadas en
alcance»: DIR dirige un area, SUP dirige un equipo— anclando el alcance con la minima
intervencion posible. Lo que la pregunta pide no se toca: los C3 (frecuencia y formato de
seguimiento; plazo y numero de personas a cargo) y los C4 quedan identicos.

Es una interpretacion nuestra, no una instruccion literal, y queda anotada como tal en la
hoja «Correcciones» del libro. Si la clienta manda otra redaccion, gana la suya.

Se ejecuta una sola vez y es idempotente: si ya esta aplicado, avisa y no toca nada.

    python3 scripts/ajusta-sup-r14-r15.py
"""

import sys

import openpyxl

LIBRO = 'docs/insumos/CAZATALENTOS-SUP.xlsx'

CAMBIOS = {
    'R14': (
        'Dame un ejemplo de una persona que producía muy bien, otra que estaba creciendo y '
        'otra que era deficiente. ¿Qué hiciste diferente con cada una y por qué?',
        'Dame un ejemplo, dentro de tu equipo, de una persona que producía muy bien, otra que '
        'estaba creciendo y otra que era deficiente. ¿Qué hiciste diferente con cada una y por qué?',
    ),
    'R15': (
        'Una persona de bajo rendimiento que gestionaste: ¿qué detectaste, qué hiciste, cuánto '
        'tomó y cómo terminó? ¿Cuánta gente tenías a cargo entonces?',
        'Una persona de tu equipo con bajo rendimiento que gestionaste: ¿qué detectaste, qué '
        'hiciste, cuánto tomó y cómo terminó? ¿Cuánta gente tenías a cargo entonces?',
    ),
}

PORQUE = (
    'INTERPRETACIÓN NUESTRA, no instrucción literal de la clienta. El .md (parte 4) pide «tu '
    'equipo» en vez de «tu área», pero ninguna de las dos preguntas contenía la palabra «área»: '
    'la instrucción parece escrita contra un borrador anterior. Se aplicó su intención declarada '
    '(«las mismas preguntas de DIR, ajustadas en alcance»: DIR dirige un área, SUP un equipo) '
    'anclando el alcance con la mínima intervención. Lo que la pregunta pide no cambia: los C3 '
    '(frecuencia y formato de seguimiento; plazo y número de personas a cargo) y los C4 quedan '
    'idénticos. SI ELLA PREFIERE OTRA REDACCIÓN, GANA LA SUYA: cambiar un enunciado después de '
    'publicar invalida la comparación con quien ya respondió.'
)


def main():
    libro = openpyxl.load_workbook(LIBRO)
    hoja = libro['Prueba RENASER']

    tocadas = {}
    for fila in hoja.iter_rows(min_row=5):
        codigo = fila[0].value
        if codigo not in CAMBIOS:
            continue
        antes, despues = CAMBIOS[codigo]
        if fila[2].value == despues:
            print('%s ya estaba ajustada; no se toca nada.' % codigo)
            return 0
        if fila[2].value != antes:
            print('%s no tiene el texto esperado. Abortado sin escribir.\n  hay: %r'
                  % (codigo, fila[2].value), file=sys.stderr)
            return 1
        fila[2].value = despues
        tocadas[codigo] = fila[2].coordinate

    if len(tocadas) != 2:
        print('Se esperaban R14 y R15; se encontraron %s. Abortado.' % sorted(tocadas),
              file=sys.stderr)
        return 1

    # La hoja «Correcciones» traía esto como pendiente. Pasa a ser un cambio aplicado.
    correcciones = libro['Correcciones']
    anotada = False
    for fila in correcciones.iter_rows(min_row=5):
        if fila[1].value == 'R14 y R15':
            fila[1].value = '%s y %s' % (tocadas['R14'], tocadas['R15'])
            fila[2].value = 'Idénticas a DIR: ninguna mencionaba equipo ni área.'
            fila[3].value = ('R14 · «Dame un ejemplo, DENTRO DE TU EQUIPO, de una persona…»  ·  '
                             'R15 · «Una persona DE TU EQUIPO con bajo rendimiento que gestionaste…»')
            fila[4].value = PORQUE
            anotada = True
    if not anotada:
        print('No se encontró la fila «R14 y R15» en la hoja Correcciones. Abortado.',
              file=sys.stderr)
        return 1

    libro.save(LIBRO)
    print('SUP ajustado · R14 en %s · R15 en %s' % (tocadas['R14'], tocadas['R15']))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
