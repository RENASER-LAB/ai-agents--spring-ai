#!/usr/bin/env python3
"""
La lista de los mejores de cada vacante, en Excel, para pasarsela a Recursos Humanos.

POR QUE EXISTE
--------------
El panel enseña el ranking en pantalla, y eso sirve para mirar. Pero quien llama por
telefono a veinte personas necesita el telefono, el correo y la nota en una hoja que pueda
ordenar, marcar y compartir, no una pantalla que hay que ir bajando.

DE DONDE SALE CADA COLUMNA
--------------------------
Todo del mismo sitio que el panel: `GET /api/v1/panel/vacantes/{id}/ranking`. No hay ningun
calculo aqui, asi que la hoja no puede decir algo distinto de lo que dice el sistema.

El correo y el telefono salen de `datos`, que es lo que el agente leyo del propio curriculum.
NO del correo de la cuenta: en esta convocatoria ese se lo invento el cargador y no existe.

EL ORDEN ES EL DEL PANEL, y no la nota pura: primero el grupo de prioridad, y dentro de el la
nota. Se pone el grupo en una columna para que se vea por que el 9 tiene mas nota que el 8.

USO
---
    python scripts/lista-para-rrhh.py                       # los 20 primeros de cada vacante
    python scripts/lista-para-rrhh.py --cuantos 30
    python scripts/lista-para-rrhh.py --salida "C:/ruta/archivo.xlsx"
"""
import argparse
import json
import sys
import urllib.error
import urllib.request

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BACKEND = "https://18-204-177-210.nip.io"
VACANTES = [(15, "Arquitecto"), (16, "Ingeniero Civil")]

# El semaforo del propio sistema, para que la hoja se lea igual que el panel.
COLOR = {
    "ALTA": "2F6E51",
    "POTENCIAL_CON_RIESGO": "9B6A22",
    "NO_PRIORIZADO": "70706B",
    "INCOMPATIBLE": "A43B36",
}
LEGIBLE = {
    "ALTA": "Alta",
    "POTENCIAL_CON_RIESGO": "Potencial con riesgo",
    "NO_PRIORIZADO": "No priorizado",
    "INCOMPATIBLE": "Incompatible",
}

COLUMNAS = [
    ("#", 5), ("Candidato", 34), ("Nota", 8), ("Prioridad", 21),
    ("Teléfono", 15), ("Correo", 34), ("Último puesto", 28),
    ("Años exp.", 10), ("Estado en el sistema", 26), ("ID", 7),
]


def llamar(metodo, ruta, token=None, cuerpo=None):
    datos = json.dumps(cuerpo).encode() if cuerpo is not None else None
    peticion = urllib.request.Request(BACKEND + ruta, data=datos, method=metodo)
    if datos is not None:
        peticion.add_header("Content-Type", "application/json")
    if token:
        peticion.add_header("Authorization", "Bearer " + token)
    try:
        with urllib.request.urlopen(peticion, timeout=60) as respuesta:
            texto = respuesta.read().decode()
            return respuesta.status, (json.loads(texto) if texto.strip() else None)
    except urllib.error.HTTPError as error:
        return error.code, error.read().decode()[:300]


def hoja_de(libro, titulo, filas, cuantos):
    h = libro.create_sheet(titulo)

    for indice, (nombre, ancho) in enumerate(COLUMNAS, 1):
        celda = h.cell(row=1, column=indice, value=nombre)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="0D0D0D")
        celda.alignment = Alignment(vertical="center")
        h.column_dimensions[get_column_letter(indice)].width = ancho
    h.freeze_panes = "A2"

    for numero, fila in enumerate(filas[:cuantos], 1):
        datos = fila.get("datos") or {}
        meses = datos.get("experienciaMesesTotal")
        grupo = fila.get("grupoPrioridad")
        valores = [
            numero,
            fila.get("candidato") or datos.get("nombre") or "",
            fila.get("notaEtapa"),
            LEGIBLE.get(grupo, grupo or ""),
            # Como texto: un numero de celular empieza por 9 y Excel se comeria el cero
            # de los fijos y lo pondria en notacion cientifica si fuera numero.
            str(datos.get("telefono") or ""),
            datos.get("email") or "",
            datos.get("ultimoPuesto") or "",
            round(meses / 12, 1) if isinstance(meses, (int, float)) else "",
            fila.get("estadoNombre") or fila.get("estado") or "",
            fila.get("postulacionId"),
        ]
        for indice, valor in enumerate(valores, 1):
            celda = h.cell(row=numero + 1, column=indice, value=valor)
            if indice == 4 and grupo in COLOR:
                celda.font = Font(color=COLOR[grupo], bold=True)
            if indice in (5, 6):
                celda.alignment = Alignment(horizontal="left")

        # Sin correo no se le puede escribir. Se marca aqui para que quien llame lo sepa
        # antes de intentarlo, y no lo descubra cuando el correo rebote.
        if not (datos.get("email") or "").strip():
            for indice in range(1, len(COLUMNAS) + 1):
                h.cell(row=numero + 1, column=indice).fill = PatternFill("solid", fgColor="FBE9E7")

    return h


def main():
    opciones = argparse.ArgumentParser(description=__doc__,
                                       formatter_class=argparse.RawDescriptionHelpFormatter)
    opciones.add_argument("--cuantos", type=int, default=20)
    opciones.add_argument("--usuario", default="andy-dev")
    opciones.add_argument("--salida", default="candidatos-convocatoria.xlsx")
    args = opciones.parse_args()

    estado, sesion = llamar("POST", "/api/v1/panel/auth/dev-login",
                            cuerpo={"usuarioRenaserOsId": args.usuario})
    if estado != 200:
        print(f"No se pudo entrar al panel: {estado} {sesion}", file=sys.stderr)
        return 1
    token = sesion["token"]

    libro = Workbook()
    libro.remove(libro.active)

    for vacante_id, titulo in VACANTES:
        estado, ranking = llamar("GET", f"/api/v1/panel/vacantes/{vacante_id}/ranking", token)
        if estado != 200:
            print(f"Vacante {vacante_id}: {estado} {ranking}", file=sys.stderr)
            return 1
        filas = ranking["filas"] if isinstance(ranking, dict) and "filas" in ranking else ranking
        hoja_de(libro, titulo, filas, args.cuantos)
        sin_correo = sum(1 for f in filas[:args.cuantos]
                         if not ((f.get("datos") or {}).get("email") or "").strip())
        print(f"{titulo}: {min(args.cuantos, len(filas))} candidatos"
              + (f" · {sin_correo} sin correo (fila en rojo claro)" if sin_correo else ""))

    libro.save(args.salida)
    print(f"\nGuardado en {args.salida}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
